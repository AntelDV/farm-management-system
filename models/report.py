from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ReportGenerator:
    @staticmethod
    def generate_excel_report(data, report_type, filename=None):
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bao_cao_{report_type}_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        # Tiêu đề báo cáo
        title_map = {
            'crops': 'BÁO CÁO CÂY TRỒNG',
            'animals': 'BÁO CÁO VẬT NUÔI',
            'activities': 'BÁO CÁO HOẠT ĐỘNG',
            'finance': 'BÁO CÁO TÀI CHÍNH',
            'inventory': 'BÁO CÁO KHO'
        }
        ws.title = title_map.get(report_type, "BÁO CÁO")
        
        # Định dạng
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Xác định cột dựa trên loại báo cáo
        if report_type == 'crops':
            headers = ["STT", "Tên cây", "Loại", "Ngày trồng", "Diện tích (ha)", "Trạng thái"]
            columns = ["id", "name", "type", "planting_date", "area", "status"]
            col_widths = [8, 25, 20, 15, 15, 20]
        elif report_type == 'animals':
            headers = ["STT", "Tên vật nuôi", "Loại", "Ngày nhập", "Số lượng", "Trạng thái"]
            columns = ["id", "name", "type", "entry_date", "quantity", "status"]
            col_widths = [8, 25, 20, 15, 15, 20]
        elif report_type == 'activities':
            headers = ["STT", "Tên hoạt động", "Loại", "Ngày", "Người phụ trách", "Trạng thái"]
            columns = ["id", "name", "type", "date", "responsible", "status"]
            col_widths = [8, 25, 20, 15, 20, 15]
        elif report_type == 'finance':
            headers = ["STT", "Ngày", "Loại", "Danh mục", "Số tiền", "Mô tả"]
            columns = ["id", "date", "type", "category", "amount", "description"]
            col_widths = [8, 15, 10, 20, 15, 30]
        elif report_type == 'inventory':
            headers = ["STT", "Tên vật tư", "Số lượng", "Đơn vị", "Đơn giá", "Tổng giá trị"]
            columns = ["id", "name", "quantity", "unit", "unit_price", ""]
            col_widths = [8, 25, 15, 10, 15, 15]
        
        # Tiêu đề báo cáo
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = title_map.get(report_type, "BÁO CÁO")
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Header
        for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Dữ liệu
        for row_num, item in enumerate(data, 4):
            for col_num, column in enumerate(columns, 1):
                if column == "" and report_type == "inventory":
                    value = item["current_quantity"] * item["unit_price"]
                else:
                    value = item.get(column, "")
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                
                if isinstance(value, (int, float)) and column != "id":
                    cell.number_format = '#,##0'
        
        # Lưu file
        if not os.path.exists('reports'):
            os.makedirs('reports')
        
        filepath = os.path.join('reports', filename)
        wb.save(filepath)
        
        return filepath